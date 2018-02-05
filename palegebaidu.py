#-*- coding:utf-8 -*-
from selenium import webdriver
import time
import os

from selenium.webdriver import ActionChains

from openpyxl import load_workbook

chromedriver = "C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe"
os.environ["webdriver.chrome.driver"] = chromedriver
driver =  webdriver.Chrome(chromedriver)
driver.maximize_window()

workbook = load_workbook("a.xlsx") #excel文件
worksheet1 = workbook.worksheets[0]
worksheet2 = workbook.worksheets[1]
rowlength = len(list(worksheet1.rows))


#driver.get("http://manage.sxtzxm.gov.cn/tzxmapp/controller.do?location=%2Ftzxmapp%2Fpages%2Fconfigure%2FcatalogAndItems%2FitemPowerList_pz.jsp")
driver.get("http://61.185.238.209:8091/tzxmspall/tzxmapp/pages/configure/catalogAndItems/itemPowerList_pz.jsp")
time.sleep(1)                            #让操作稍微停一下
#找到输入账号的框，并自动输入账号 这里要替换为你的登录账号
driver.find_element_by_id('j_username').send_keys('test')
time.sleep(1)
#密码，这里要替换为你的密码
driver.find_element_by_id('j_password').send_keys('2')
#driver.find_element_by_id('j_password').send_keys('Q!W@e3r4')
time.sleep(1)
#找到登录按钮，并点击
driver.find_element_by_class_name('loginbtn').click()
time.sleep(5)

#driver.get("http://manage.sxtzxm.gov.cn/tzxmapp/controller.do?location=%2Ftzxmapp%2Fpages%2Fconfigure%2FcatalogAndItems%2FitemPowerList_pz.jsp")
driver.get("http://61.185.238.209:8091/tzxmspall/tzxmapp/pages/configure/catalogAndItems/itemPowerList_pz.jsp")
time.sleep(3)



for rm in range(1,rowlength):
    sx_code = worksheet1.cell(row=rm,column=1).value
    ry_name = worksheet1.cell(row=rm,column=2).value

    driver.find_element_by_id("item_code").find_element_by_class_name("u-form-textbox-input").send_keys(sx_code)
    # 搜索事项编码
    driver.find_element_by_id("unieap_form_Button_0").find_element_by_class_name("u-form-btn").click()
    time.sleep(1)
    driver.find_element_by_id("gridResultDb").find_element_by_id("gridResultDb_select_1").click()
    time.sleep(1)
    driver.find_element_by_id("btnMaterial").click()

    rightlilist = driver.find_element_by_id("right_items").find_elements_by_tag_name("li")
    for i in range(0,len(rightlilist)):
        aa = rightlilist[i].get_attribute("id")
        time.sleep(1)
        rightdata = driver.find_element_by_id(aa)
        ActionChains(driver).double_click(rightdata).perform()


    time.sleep(1)
    flag = "false"
    leftlilist = driver.find_element_by_id("left_items").find_elements_by_tag_name("li")
    for j in range(0, len(leftlilist)):
        listname = leftlilist[j].find_elements_by_tag_name("div")
        enselectname = listname[1].text
        if ry_name == enselectname:
            bb = leftlilist[j].get_attribute("id")
            time.sleep(1)
            leftdata = driver.find_element_by_id(bb)
            ActionChains(driver).double_click(leftdata).perform()
            flag = "true"

    if flag == "false":
        wkrow = len(list(worksheet2.rows))+1
        worksheet2.cell(row=wkrow,column=1).value = sx_code
        worksheet2.cell(row=wkrow,column=2).value = ry_name

    time.sleep(1)
    driver.find_element_by_id("btnBack").click()

workbook.save("a.xlsx")
