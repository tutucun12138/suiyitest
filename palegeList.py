from selenium import webdriver
import time
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import os

from selenium.webdriver import ActionChains

chromedriver = "C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe"
os.environ["webdriver.chrome.driver"] = chromedriver
driver =  webdriver.Chrome(chromedriver)
driver.maximize_window()

driver.get("http://manage.sxtzxm.gov.cn/tzxmapp/controller.do?location=%2Ftzxmapp%2Fpages%2Fconfigure%2FcatalogAndItems%2FitemPowerList_pz.jsp")
time.sleep(1)                            #让操作稍微停一下
#找到输入账号的框，并自动输入账号 这里要替换为你的登录账号
driver.find_element_by_id('j_username').send_keys('test')
time.sleep(1)
#密码，这里要替换为你的密码
driver.find_element_by_id('j_password').send_keys('Q!W@e3r4')
time.sleep(1)
#找到登录按钮，并点击
driver.find_element_by_class_name('loginbtn').click()
time.sleep(5)

driver.get("http://manage.sxtzxm.gov.cn/tzxmapp/controller.do?location=%2Ftzxmapp%2Fpages%2Fconfigure%2FcatalogAndItems%2FitemPowerList_pz.jsp")
time.sleep(5)

workbook = load_workbook("a.xlsx")
worksheet = workbook.worksheets[0] #excel准备
areacodelist = ["610300"]
for i in range(0,len(areacodelist)):
    driver.find_element_by_id("item_code").find_element_by_class_name("u-form-textbox-input").send_keys(areacodelist[i])
    driver.find_element_by_id("unieap_form_Button_0").find_element_by_class_name("u-form-btn").click()
    time.sleep(1)
    pagenum = driver.find_element_by_id("gridResultDb_paginate").find_elements("xpath","//span/a")
    countpage = pagenum[len(pagenum)-1].text  #总页数
    for j in range(0,int(countpage)):
        trlist = driver.find_element_by_id("gridResultDb").find_elements_by_tag_name("tr")
        for k in range(0,len(trlist)):
            workrow = worksheet.rows
            tdlist = trlist[k].find_elements_by_tag_name("td")
            print(tdlist)
            #worksheet.cell(workrow+1,1).value = tdlist[1]
            #worksheet.cell(workrow+1,2).value = tdlist[2]
            #worksheet.cell(workrow+1,3).value = tdlist[3]
        time.sleep(1)
        driver.find_element_by_id("gridResultDb_next").click()
workbook.save("a.xlsx")



